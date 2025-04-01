import calendar
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.drawing.image import Image

# Constants
DAYS_OF_WEEK = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
ATTENDANCE_OPTIONS = ["P", "L", "E", "A","N"]
REMARKS = {90: "Excellent", 75: "Good", 50: "Average", 0: "Poor"}

# Function to create attendance sheet for a month
def create_month_sheet(workbook, month, year):
    sheet = workbook.create_sheet(title=month)
    # sheet.freeze_panes = "A15"  # Freeze rows up to row 14

    # Set the header row
    sheet["D17"] = "Name"
    sheet['C17'] = 'No.'
    current_date = datetime(year, list(calendar.month_name).index(month), 1)
    col = 5
    total_days = 0

    while current_date.month == list(calendar.month_name).index(month):
        if current_date.strftime("%A") in DAYS_OF_WEEK:
            total_days += 1
            day_cell = sheet.cell(row=17, column=col)
            date_cell = sheet.cell(row=18, column=col)
            day_cell.value = current_date.strftime("%a")
            date_cell.value = int(current_date.strftime("%d"))
            day_cell.alignment = date_cell.alignment = Alignment(horizontal="center")
            col += 1
        current_date += timedelta(days=1)

    # Add average attendance and remarks columns
    for K in ATTENDANCE_OPTIONS:
        k = col + ATTENDANCE_OPTIONS.index(K) + 1
        sheet.cell(row=17, column=k).value = K

    attendance_col = col + 7
    avg_col = col + 8
    remarks_col = col + 9
    sheet.cell(row=17, column=attendance_col).value = "Attendance"
    sheet.cell(row=17, column=avg_col).value = "Average Attendance (%)"
    sheet.cell(row=17, column=remarks_col).value = "Remarks"

    return sheet, avg_col, remarks_col, total_days

def col_let(number: int): # Returns the column letter for an number
    return get_column_letter(number)

# Function to calculate attendance summary and add dropdown options
def add_summary(workbook, workers, total_days_dict):

    for sheet in workbook.worksheets:
        total_days= total_days_dict[sheet.title]
        avg_col = total_days + 13
        remarks_col = avg_col + 1

        # Create a dictionary using dictionary comprehension
        dict_AO_pos = {ATTENDANCE_OPTIONS[i]: col_let(total_days+6+i) for i in range(len(ATTENDANCE_OPTIONS))}

        dd_opts = list(dict_AO_pos.keys()) # Dropdown options
        # Create a data validation for the dropdown
        dv = DataValidation(type="list", formula1=f'"{",".join(dd_opts)}"', showDropDown=True)

        # Add the data validation to each sheet
        sheet.add_data_validation(dv)

        for row in range(19, 19 + len(workers)):
            for column in range(total_days + 6, total_days + 11):
                position = column - total_days - 6
                opt_list = list(dict_AO_pos.values()) # Get a list of columns for attendance summary
                sheet.cell(row=row, column=column ).value = f"=COUNTIF(E{row}:{col_let(total_days+4)}{row},${opt_list[position]}$17)"
                dv.add(f"E{row}:{col_let(total_days+4)}{row}")

            sheet[
                f"{col_let(total_days + 12)}{row}"].value = f"=SUM({col_let(total_days + 6)}{row},{col_let(total_days + 7)}{row})"

            sheet[
                f"{col_let(total_days + 13)}{row}"].value = f"=ROUND({col_let(total_days + 12)}{row}*{100}/({col_let(total_days + 6)}{row}+{col_let(total_days + 7)}{row}+{col_let(total_days + 8)}{row}+{col_let(total_days + 9)}{row}),2)"

            avg_cell = sheet.cell(row=row, column= avg_col)
            remark_cell = sheet.cell(row=row, column= remarks_col)
            remark_cell.value = (f'=IF({avg_cell.coordinate}>=85.00,"Excellent",IF({avg_cell.coordinate}>=75.00,"Good",'
                                f'IF({avg_cell.coordinate}>=60.00,"Average","Poor")))')

        # Add rows for total attendance statistics for each day

        sheet.cell(row=95, column=4).value = "TOTAL PUNCTUAL"
        sheet.cell(row=96, column=4).value = "TOTAL LATENESS"
        sheet.cell(row=97, column=4).value = "TOTAL EXCUSES"
        sheet.cell(row=98, column=4).value = "TOTAL ABSENTS"
        sheet.cell(row=99, column=4).value = "NO MEETING"
        sheet.cell(row=100, column=4).value = "TOTAL ATTENDANCE"

        for col in range(5, total_days + 5):
            col_letter = get_column_letter(col)  # Convert column index to letter

            sheet.cell(row=95, column=col).value = f'=COUNTIF({col_letter}{19}:{col_letter}{len(workers) + 19}, "P")'
            sheet.cell(row=96, column=col).value = f'=COUNTIF({col_letter}{19}:{col_letter}{len(workers) + 19}, "L")'
            sheet.cell(row=97, column=col).value = f'=COUNTIF({col_letter}{19}:{col_letter}{len(workers) + 19}, "E")'
            sheet.cell(row=98, column=col).value = f'=COUNTIF({col_letter}{19}:{col_letter}{len(workers) + 19}, "A")'
            sheet.cell(row=99, column=col).value = f'=COUNTIF({col_letter}{19}:{col_letter}{len(workers) + 19}, "N")'
            sheet.cell(row=100, column=col).value = (
                f'=COUNTIF({col_letter}{19}:{col_letter}{len(workers) + 19}, "P") +'
                f'COUNTIF({col_letter}{19}:{col_letter}{len(workers) + 19}, "L")')


def handle_division_by_zero_errors(workbook,total_days_dict):
    for sheet in workbook.worksheets:
        col_num = total_days_dict[sheet.title] + 13
        for row in sheet[f"{col_let(col_num)}19:{col_let(col_num)}93"]:
            for cell in row:
                # Check if the cell contains an error value (string starting with '#')
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith("#"):
                    cell.value = 0


# Function to plot charts
def plot_charts(workbook, workers, total_days_dict):
    for sheet in workbook.worksheets:
        total_days = total_days_dict[sheet.title]
        avg_col = total_days + 13

        # Add a chart_1 for individual attendance
        chart_1 = BarChart()
        chart_1.title = f"Individual Average Daily Attendance ({sheet.title})"
        chart_1.x_axis.title = "Attendees"
        chart_1.y_axis.title = "Days_present"
        data_1 = Reference(sheet, min_col=avg_col, min_row=18, max_row=18 + len(workers))
        chart_1.add_data(data_1, titles_from_data=True)
        sheet.add_chart(chart_1, f"E2")

        # Add line chart_1 for collective attendance
        chart_2 = BarChart()
        chart_2.title = "Total Daily Attendance"
        chart_2.x_axis.title = "Dates"
        chart_2.y_axis.title = "Total Attendance"
        # Define the data for the chart_1
        data_2 = Reference(sheet,
                         min_col=4,  # Starting from column D
                         min_row=100,
                         max_col=total_days + 4,
                         max_row=100)  # Only the 42nd row

        # Define the x-axis labels (dates) from row 1 (assuming headers are in row 1)
        dates = Reference(sheet,
                          min_col=4,
                          min_row=18,
                          max_col=total_days + 4,  # Same number as max_col data
                          max_row=18)  # Assume headers are in row
        chart_2.add_data(data_2, from_rows=False)
        chart_2.set_categories(dates)
        sheet.add_chart(chart_2, f"Q2")

def color_fill(colour: str):
    return PatternFill(start_color=f'{colour}', end_color=f'{colour}', fill_type='solid')


def format_sheet(workbook, workers, total_days_dict):
    for sheet in workbook.worksheets:
        total_days = total_days_dict[sheet.title]

        end_col = col_let(total_days+16)
        # Formating row 1 to row 16 and column A
        cell_row_range = sheet[f'A4:{end_col}16']
        for row in cell_row_range:
            for cell_row in row:
                cell_row.fill = color_fill('f1f3fd')

        # Format columns A and B and the 3rd colum after the remark colum
        for row in range(17, 101):  # 101 is excluded
            sheet[f'A{row}'].fill = color_fill('f1f3fd') # For column A
            sheet[f'B{row}'].fill = color_fill('fdfefe') # For column B

        # General background, text and border formatting
        cell_range = sheet[f'C17:{end_col}100']
        font_style = Font(name='Arial', size=9, bold=True, color='000000')  # Black, bold font
        border_style = Border(left=Side(style='none'),
                             right=Side(style='none'),
                             top=Side(style='hair', color='00CCCCCC'),
                            bottom=Side(style='hair', color='00CCCCCC'))  # Thin border
        alignment_style = Alignment(horizontal='center', vertical='center')

        for row in cell_range:
            for cell in row:
                cell.font = font_style
                cell.fill = color_fill("fdfefe")
                cell.border = border_style
                cell.alignment = alignment_style

        # formatting for first 3 rows
        row_1_3 = sheet[f'A1:{col_let(total_days+17)}3']

        for row in  row_1_3:
            for cell_1_3 in row:
                cell_1_3.fill = color_fill('7aa0f8')
                cell_1_3.alignment = alignment_style

        # Color fill for the remaining of the rows in 3rd column after remark column
        for row in sheet[f'{col_let(total_days+17)}4:{col_let(total_days+17)}100']:
            for cell in row:
                cell.fill = color_fill('f1f3fd')

        # Create conditional fill pattern for cells with value P
        for column in range(5, total_days + 5):
            col_letter = col_let(column)
            cell_range = f"{col_letter}19:{col_letter}{18+len(workers)}"
            # Add conditional formatting rule
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=['"P"'], fill=color_fill('E5EFF8'))
            )
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=['"E"'], fill=color_fill('F7F2E2'))
            )
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=['"L"'], fill=color_fill('EFF7EA'))
            )

        # Add conditional formating to the remarks column
        ave_col_let = col_let(total_days + 14)
        remarks_color= {'28a745': "Excellent", '86c3e0': "Good", 'ffc107': "Average", 'dc3545': "Poor"}
        cell_range = f"{ave_col_let}19:{ave_col_let}{19 + len(workers)}"
        for colour, remark in remarks_color.items():
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=[f'"{remark}"'], fill=color_fill(colour)))

        # Format the day and date rows
        omit = list([col_let(total_days + x) for x in range(5, 12, 6)])
        omit_cells = {f'{col}17' for col in omit}.union({f'{col}18' for col in omit}) # Create a set of omitted cell
        for row in sheet[f'C17:{ave_col_let}18']:
            for cell in row:
                if cell.coordinate not in omit_cells:
                    cell.fill = color_fill('7ab0fa')
                    cell.font = Font(name='Arial', size=8, bold=True, color='ffffff')  # White font

        # color formatting for columns where the date is Monday,
        wed_thu = ['Wed', 'Thu'] # Wednesday or Thursday
        start_col, end_col = 3, 3 + total_days
        for row in range(19, 101):
            for col in range(start_col, end_col + 1):
                column_letter = col_let(col)
                if sheet[f'{column_letter}17'].value == "Mon":
                    cells = sheet.cell(row=row, column=col)
                    cells.fill = color_fill('DEE4F2') #
                elif sheet[f'{column_letter}17'].value in wed_thu:
                    cells = sheet.cell(row=row, column=col)
                    cells.fill = color_fill('ededee')

        # Formatting for the totals
        for column in range(total_days + 6, total_days + 11, 2):
            for row in range(19, 101):
                cells = sheet.cell(row=row, column=column)
                if sheet[f'{col_let(column)}17'].value in ['P', 'E', 'N']:
                    cells.fill = color_fill('E6DED2')

        # Formatting for the last row
        for row in sheet[f'{col_let(1)}101:{col_let(total_days + 17)}101']:
            for cell in row:
                cell.fill = color_fill('f1f3fd')

        # Pattern fill for the remaining (Unused cells)
        for row in range(1, 101):
            for cols in range(total_days + 18, 251):
                cell_un = sheet.cell(row=row, column=cols)
                cell_un.fill = color_fill('CCE8CF')

        for row in range(102, 321):
            for col in range(1, 231):
                cell_un = sheet.cell(row=row, column=col)
                cell_un.fill = color_fill('CCE8CF')

def format_cell_size(workbook, total_days_dict):
    for sheet in workbook.worksheets:
        td = total_days_dict[sheet.title]
        omit_cols = list([col_let(td + x) for x in range(5, 12, 6)])

        sheet.row_dimensions[2].height = 25
        sheet.row_dimensions[16].height = 20
        sheet.column_dimensions['A'].width = 4
        sheet.column_dimensions['B'].width = 6
        sheet.column_dimensions['D'].width = 29
        sheet.column_dimensions[f'{col_let(td+5)}'].width = 3
        sheet.column_dimensions[f'{col_let(td +11 )}'].width = 3
        sheet.column_dimensions[f'{col_let(td + 17)}'].width = 4

        for col in range(5, td + 12):
            if col_let(col) not in omit_cols:
                sheet.column_dimensions[col_let(col)].width = 5

# Add other write ups
def writings(workbook, total_days_dict):
    cell_alignment = Alignment(horizontal='center', vertical='center')
    for sheet in workbook.worksheets:
        td = total_days_dict[sheet.title]
        col_k = col_let(td + 15)
        col_v = col_let(td + 16)
        rowed = 20

        sheet['H2'].font = Font(name='Arial', size=18, bold=True, color='ffffff')  # White, bold font
        sheet['H2'].alignment = cell_alignment
        sheet['H2'].value = 'Gocity Group Developers / Engineers Attendance Record'
        sheet[f'{col_k}5'].value = 'P ='
        sheet[f'{col_v}5'].value = 'Punctual'
        sheet[f'{col_k}6'].value = 'E ='
        sheet[f'{col_v}6'].value = 'Excused'
        sheet[f'{col_k}7'].value = 'L ='
        sheet[f'{col_v}7'].value = 'Late'
        sheet[f'{col_k}8'].value = 'A ='
        sheet[f'{col_v}8'].value = 'Absent'
        sheet[f'{col_k}9'].value = 'N ='
        sheet[f'{col_v}9'].value = 'No Meeting'
        sheet[f'{col_k}11'].value = 'Working Days ='
        sheet[f'{col_v}11'].value = td
        sheet[f'{col_k}12'].value = 'Days Covered ='
        sheet[f'{col_v}12'].value = f"=SUM({col_let(td + 6)}{rowed}:{col_let(td + 9)}{rowed})"
        sheet[f'{col_k}15'].value = 'Month:'
        sheet[f'{col_v}15'].value = f'{sheet.title}'
        sheet[f'{col_k}16'].value = 'TODAY:'
        sheet[f'{col_let(td+16)}16'].value = "=TODAY()"

        for row in sheet[f'{col_k}4:{col_k}16']:
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')


def add_images(workbook, cell_location, image_path, width= 100, height= 100):
    img = Image(image_path)
    img.width = width
    img.height = height
    for sheet in workbook.worksheets:
        sheet.add_image(img, cell_location)
        # cell = sheet[cell_location]  # Get the cell object
        # column_letter = cell_location[0]  # Extract the column letter (e.g., "A")
        # sheet.column_dimensions[column_letter].width = 20  # Adjust as needed
        # sheet.row_dimensions[cell.row].height = 70  # Adjust as needed


# Main function
def main():
    workbook = Workbook()
    year = datetime.now().year
    workbook.remove(workbook.active)  # Remove default sheet

    # Example workers list
    workers = ['Sheriff Akatugba', 'Bassey Nton Nton', 'Denis Esikpong', 'Obiora Anointing Chimnadindu',
               'Maurice Tchouncha', 'Nnamso Glory', 'Joseph Onyinye', 'Agbor Wilmont', 'Emmamuel Aniefiok',
               'Ofonime Ufot', 'Blessing Edet', 'Victor Emordi', 'Christopher Monday', 'Chris Okoh', "", "", "", "", "", "",  "", "", "", "", "", ""]

    total_days_dict = {}

    # Create sheets for each month
    for month in list(calendar.month_name)[1:]:
        sheet, avlg_col, remarks_col, total_days = create_month_sheet(workbook, month, year)
        total_days_dict[month] = total_days

        # Add worker names
        for i, worker in enumerate(workers, start=19):
            sheet.cell(row=i, column=4).value = worker

    # Add attendance summary
    add_summary(workbook, workers, total_days_dict)

    # Add error handler for Division by Zero Error
    handle_division_by_zero_errors(workbook, total_days_dict)

    # Add Charts
    plot_charts(workbook, workers, total_days_dict)

    # Add formatting
    format_sheet(workbook, workers, total_days_dict)

    # Add size formatting
    format_cell_size(workbook, total_days_dict)

    # Adding other Writings
    writings(workbook, total_days_dict)

    # Add Logo
    go_logo = "GOCITY_1.png"
    add_images(workbook, 'A2',go_logo,)

    # Save the workbook
    workbook.save("GOCITY_TECH_ATTENDANCE_REGISTER.xlsx")

if __name__ == "__main__":
    main()
