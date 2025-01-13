import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import calendar

# Create a new workbook and remove the default sheet
wb = Workbook()
wb.remove(wb.active)

# Worker names
workers = ['Ashani Akatugba', 'Basey Nton N,', 'Anointing Obiora', 'Denis Esikpong',
           'Maurice Tchouncha', 'Nnamso Glory', 'Joseph Onyinye', 'Emmamuel Aniefiok',
           'Ofonime Emmanuel', 'Blessing Edet','Victor Emordi', 'Chris Okoh', 'Christopher']


def create_month_sheet(month, year=2025):
    days = calendar.monthrange(year, month)[1]
    start_date = datetime(year, month, 1)

    # Create a new sheet for the month
    month_name = calendar.month_name[month]
    ws = wb.create_sheet(month_name)

    # Fill the first row with dates and weekdays, skip weekends
    dates = []
    for day in range(1, days + 1):
        date = datetime(year, month, day)
        if date.weekday() < 5:  # Monday to Friday are 0-4
            dates.append(date.strftime("%Y-%m-%d (%a)"))

    # Preparing data frame
    data = {
        'Names': workers
    }
    for date in dates:
        data[date] = ['' for _ in workers]

    df = pd.DataFrame(data)

    # Load data into the worksheet, starting from row 15
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 15):
        ws.append(row)

    # Calculate averages and remarks
    for row in range(16, 16 + len(workers)):
        cell = ws.cell(row, len(dates) + 2)
        cell.value = f'=AVERAGE({ws.cell(row, 2).coordinate}:{ws.cell(row, 1 + len(dates)).coordinate})'
        remark_cell = ws.cell(row, len(dates) + 3)
        remark_cell.value = f'=IF({cell.coordinate}>0.9,"Excellent",IF({cell.coordinate}>0.75,"Good",IF({cell.coordinate}>0.5,"Average","Poor")))'

    # Chart for the average daily attendance
    chart = BarChart()
    chart.title = "Average Daily Attendance"
    data = Reference(ws, min_col=2, min_row=15, max_row=15 + len(workers), max_col=1 + len(dates))
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, f"A1")


# Generate sheets for each month of the year 2025
for month in range(1, 13):
    create_month_sheet(month)

# Save the workbook
wb.save('Worker_Training_Tracker.xlsx')

