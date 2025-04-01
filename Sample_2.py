import calendar
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule

# Constants
DAYS_OF_WEEK = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
ATTENDANCE_OPTIONS = ["P", "L", "E", "A","N"]
REMARKS = {90: "Excellent", 75: "Good", 50: "Average", 0: "Poor"}

# Function to create attendance sheet for a month
def create_month_sheet(workbook, month, year):
    sheet = workbook.create_sheet(title=month)
    # sheet.freeze_panes = "A15"  # Freeze rows up to row 14

    # Set the header row
    sheet["D17"] = "Name"
    current_date = datetime(year, list(calendar.month_name).index(month), 1)
    col = 5
    total_days = 0

    while current_date.month == list(calendar.month_name).index(month):
        if current_date.strftime("%A") in DAYS_OF_WEEK:
            total_days += 1
            day_cell = sheet.cell(row=17, column=col)
            date_cell = sheet.cell(row=18, column=col)
            day_cell.value = current_date.strftime("%a")
            date_cell.value = int(current_date.strftime("%d"))
            day_cell.alignment = date_cell.alignment = Alignment(horizontal="center")
            col += 1
        current_date += timedelta(days=1)

    # Add average attendance and remarks columns
    for K in ATTENDANCE_OPTIONS:
        k = col + ATTENDANCE_OPTIONS.index(K) + 1
        sheet.cell(row=17, column=k).value = K

    attendance_col = col + 7
    avg_col = col + 8
    remarks_col = col + 9
    sheet.cell(row=17, column=attendance_col).value = "Attendance"
    sheet.cell(row=17, column=avg_col).value = "Average Attendance (%)"
    sheet.cell(row=17, column=remarks_col).value = "Remarks"

    return sheet, avg_col, remarks_col, total_days

# Function to calculate attendance summary and add dropdown options
def add_summary(workbook, workers, total_days_dict):

    for sheet in workbook.worksheets:
        total_days= total_days_dict[sheet.title]
        avg_col = total_days + 13
        remarks_col = avg_col + 1

        dict_day = {
            23: {"P": 'AC', "L": 'AD', "E": 'AE', "A": 'AF', "N": 'AG'},
            22: {"P": 'AB', "L": 'AC', "E": 'AD', "A": 'AE', "N": 'AF'},
            21: {"P": 'AA', "L": 'AB', "E": 'AC', "A": 'AD', "N": 'AE'},
            20: {"P": 'Z', "L": 'AA', "E": 'AB', "A": 'AC', "N": 'AD'}
        }

        dd_opts = list(list(dict_day.values())[0].keys())  # Dropdown options
        # Create a data validation for the dropdown
        dv = DataValidation(type="list", formula1=f'"{",".join(dd_opts)}"', showDropDown=True)

        # Add the data validation to each sheet
        sheet.add_data_validation(dv)

        for row in range(19, 19 + len(workers)):
            for column in range(total_days + 6, total_days + 11):
                position = column - total_days - 6
                if total_days == 23:
                    opt_list = list(dict_day[23].values())
                    sheet.cell(row=row, column=column ).value = f"=COUNTIF(E{row}:AA{row},${opt_list[position]}$17)"
                    sheet[f"AI{row}"].value = f"=SUM(AC{row}:AD{row})"
                    sheet[f"AJ{row}"].value = f"=ROUND(AI{row}/{0.01 * total_days}, 1)"
                    dv.add(f"E{row}:AA{row}")


                elif total_days == 22:
                    opt_list = list(dict_day[22].values())
                    sheet.cell(row=row, column=column ).value = f"=COUNTIF(E{row}:Z{row},${opt_list[position]}$17)"
                    sheet[f"AH{row}"].value = f"=SUM(AB{row}:AC{row})"
                    sheet[f"AI{row}"].value = f"=ROUND(AH{row}/{0.01 * total_days}, 1)"
                    dv.add(f"E{row}:Z{row}")

                elif total_days == 21:
                    opt_list = list(dict_day[21].values())
                    sheet.cell(row=row, column=column ).value = f"=COUNTIF(E{row}:Y{row},${opt_list[position]}$17)"
                    sheet[f"AG{row}"].value = f"=SUM(AA{row}:AB{row})"
                    sheet[f"AH{row}"].value = f"=ROUND(AG{row}/{0.01 * total_days}, 1)"
                    dv.add(f"E{row}:Y{row}")

                elif total_days == 20:
                    opt_list = list(dict_day[20].values())
                    sheet.cell(row=row, column=column ).value = f"=COUNTIF(E{row}:X{row},${opt_list[position]}$17)"
                    sheet[f"AF{row}"].value = f"=SUM(Z{row}:AA{row})"
                    sheet[f"AG{row}"].value = f"=ROUND(AF{row}/{0.01 * total_days}, 1)"
                    dv.add(f"E{row}:X{row}")

                else:
                    pass

            avg_cell = sheet.cell(row=row, column= avg_col)
            remark_cell = sheet.cell(row=row, column= remarks_col)
            remark_cell.value = (f'=IF({avg_cell.coordinate}>90.00,"Excellent",IF({avg_cell.coordinate}>75.00,"Good",'
                                f'IF({avg_cell.coordinate}>50.00,"Average","Poor")))')

        # Add a row for total attendance for each day
        sheet.cell(row=42, column=4).value = "Total Attendance"
        for col in range(5, total_days + 5):
            col_letter = get_column_letter(col)  # Convert column index to letter
            sheet.cell(row=42, column=col).value = (f'=COUNTIFS({col_letter}{19}:{col_letter}{len(workers)+19}, "P") +'
                                                    f'COUNTIFS({col_letter}{19}:{col_letter}{len(workers)+19}, "L")')


# Function to plot charts
def plot_charts(workbook, workers, total_days_dict):
    for sheet in workbook.worksheets:
        total_days = total_days_dict[sheet.title]
        avg_col = total_days + 13

        # Add a chart_1 for individual attendance
        chart_1 = BarChart()
        chart_1.title = f"Individual Average Daily Attendance ({sheet.title})"
        chart_1.x_axis.title = "Attendees"
        chart_1.y_axis.title = "Days_present"
        data_1 = Reference(sheet, min_col=avg_col, min_row=18, max_row=18 + len(workers))
        chart_1.add_data(data_1, titles_from_data=True)
        sheet.add_chart(chart_1, f"E2")

        # Add line chart_1 for collective attendance
        chart_2 = BarChart()
        chart_2.title = "Total Daily Attendance"
        chart_2.x_axis.title = "Dates"
        chart_2.y_axis.title = "Total Attendance"
        # Define the data for the chart_1
        data_2 = Reference(sheet,
                         min_col=4,  # Starting from column D
                         min_row=42,
                         max_col=total_days + 4,
                         max_row=42)  # Only the 42nd row

        # Define the x-axis labels (dates) from row 1 (assuming headers are in row 1)
        dates = Reference(sheet,
                          min_col=4,
                          min_row=18,
                          max_col=total_days + 4,  # Same number as max_col data
                          max_row=18)  # Assume headers are in row
        chart_2.add_data(data_2, from_rows=False)
        chart_2.set_categories(dates)
        sheet.add_chart(chart_2, f"Q2")

def color_fill(colour: str):
    return PatternFill(start_color=f'{colour}', end_color=f'{colour}', fill_type='solid')

def col_let(number: int): # Returns the column letter for and number
    return get_column_letter(number)

def format_sheet(workbook, workers, total_days_dict):
    for sheet in workbook.worksheets:
        total_days = total_days_dict[sheet.title]

        end_col = col_let(total_days+16)
        # Formating row 1 to row 16 and column A
        cell_row_range = sheet[f'A4:{end_col}17']
        for row in cell_row_range:
            for cell_row in row:
                cell_row.fill = color_fill('f1f3fd')

        # Format columns A,B and 3rd col after remark column
        for row in range(17, 101):  # 101 is excluded
            sheet[f'A{row}'].fill = color_fill('f1f3fd') # For column A
            sheet[f'B{row}'].fill = color_fill('fdfefe') # For column B
            sheet[f'{col_let(total_days+17)}{row}'].fill = color_fill('f1f3fd')

        # General background, text and border formatting
        cell_range = sheet[f'C17:{end_col}100']
        font_style = Font(name='Arial', size=9, bold=True, color='000000')  # Black, bold font
        border_style = Border(left=Side(style='none'),
                             right=Side(style='none'),
                             top=Side(style='hair', color='00CCCCCC'),
                            bottom=Side(style='hair', color='00CCCCCC'))  # Thin border
        alignment_style = Alignment(horizontal='center', vertical='center')

        for row in cell_range:
            for cell in row:
                cell.font = font_style
                cell.fill = color_fill("fdfefe")
                cell.border = border_style
                cell.alignment = alignment_style

        # formatting for first 3 rows
        row_1_3 = sheet[f'A1:{end_col}3']
        font_style_1_3 = Font(name='Arial', size=9, bold=True, color='ffffff')
        for row in  row_1_3:
            for cell_1_3 in row:
                cell_1_3.font = font_style_1_3
                cell_1_3.fill = color_fill('7aa0f8')
                cell_1_3.alignment = alignment_style



        # Create conditional fill pattern for cells with value P
        for column in range(5, total_days + 5):
            col_letter = col_let(column)
            cell_range = f"{col_letter}19:{col_letter}{18+len(workers)}"
            # Add conditional formatting rule
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=['"P"'], fill=color_fill('E5EFF8'))
            )
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=['"P'
                                                      'E"'], fill=color_fill('F7F2E2'))
            )
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=['"L"'], fill=color_fill('EFF7EA'))
            )

        # Add conditional formating to the remarks column
        ave_col_let = col_let(total_days + 14)
        remarks_color= {'28a745': "Excellent", '86c3e0': "Good", 'ffc107': "Average", 'dc3545': "Poor"}
        cell_range = f"{ave_col_let}19:{ave_col_let}{19 + len(workers)}"
        for colour, remark in remarks_color.items():
            sheet.conditional_formatting.add(
                cell_range,
                CellIsRule(operator='equal', formula=[f'"{remark}"'], fill=color_fill(colour)))

        # Format the day and date rows
        omit = list([col_let(total_days + x) for x in range(5, 12, 6)])
        omit_cells = {f'{col}17' for col in omit}.union({f'{col}18' for col in omit}) # Create a set of omitted cell
        for row in sheet[f'C17:{ave_col_let}18']:
            for cell in row:
                if cell.coordinate not in omit_cells:
                    cell.fill = color_fill('7ab0fa')
                    cell.font = font_style_1_3

        # color formatting for columns where the date is Monday,
        wed_thu = ['Wed', 'Thu'] # Wednesday or Thursday
        start_col, end_col = 3, 3 + total_days
        for row in range(19, 101):
            for col in range(start_col, end_col + 1):
                column_letter = col_let(col)
                if sheet[f'{column_letter}17'].value == "Mon":
                    cells = sheet.cell(row=row, column=col)
                    cells.fill = color_fill('DEE4F2') #
                elif sheet[f'{column_letter}17'].value in wed_thu:
                    cells = sheet.cell(row=row, column=col)
                    cells.fill = color_fill('ededee')

        # Formatting for the totals
        for column in range(total_days + 6, total_days + 11, 2):
            for row in range(19, 101):
                cells = sheet.cell(row=row, column=column)
                if sheet[f'{col_let(column)}17'].value in ['P', 'E', 'N']:
                    cells.fill = color_fill('E6DED2')

        # Pattern fill for the remaining (Unused cells)
        for row in range(1, 101):
            for cols in range(total_days + 18, 251):
                cell_un = sheet.cell(row=row, column=cols)
                cell_un.fill = color_fill('CCE8CF')

        for row in range(101, 251):
            for col in range(1, 251):
                cell_un = sheet.cell(row=row, column=col)
                cell_un.fill = color_fill('CCE8CF')


# Main function
def main():
    workbook = Workbook()
    year = datetime.now().year
    workbook.remove(workbook.active)  # Remove default sheet

    # Example workers list
    workers = ['Sheriff Akatugba', 'Bassey Nton Nton', 'Denis Esikpong', 'Obiora Anointing Chimnadindu',
               'Maurice Tchouncha', 'Nnamso Glory', 'Joseph Onyinye', 'Agbor Wilmont', 'Emmamuel Aniefiok',
               'Ofonime Ufot', 'Blessing Edet', 'Victor Emordi', 'Chris Okoh', 'Christopher Monday', "", "", "", "", ""]

    total_days_dict = {}

    # Create sheets for each month
    for month in list(calendar.month_name)[1:]:
        sheet, avg_col, remarks_col, total_days = create_month_sheet(workbook, month, year)
        total_days_dict[month] = total_days

        # Add worker names
        for i, worker in enumerate(workers, start=19):
            sheet.cell(row=i, column=4).value = worker

    # Add attendance summary
    add_summary(workbook, workers, total_days_dict)

    # Add Charts
    plot_charts(workbook, workers, total_days_dict)

    # Add formatting
    format_sheet(workbook, workers, total_days_dict)

    # Add size formatting
    # format_cell_size(workbook, workers,total_days_dict)

    # Save the workbook
    workbook.save("ADJ_ATTENDANCE_PROTOTYPE.xlsx")

if __name__ == "__main__":
    main()
