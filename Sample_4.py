from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Reduce redundant DataValidation creation by applying it in batches
def apply_data_validation(sheet, start_row, end_row, start_col, end_col):
    dv = DataValidation(type="list", formula1='"P,L,E,A,N"', showDropDown=True)
    sheet.add_data_validation(dv)
    for row in range(start_row, end_row):
        dv.add(f'{get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row}')

# Optimize adding formulas by reducing redundant operations
def add_formulas(sheet, total_days, workers):
    # Add formulas for attendance count and average attendance
    for row in range(19, 19 + len(workers)):
        for col in range(5, total_days + 5):
            # Only add necessary formulas once per cell
            sheet[f'{get_column_letter(col)}{row}'].value = f'=COUNTIF(E{row}:{get_column_letter(total_days+4)}{row}, "P")'

        sheet[f'{get_column_letter(total_days + 12)}{row}'].value = f'=SUM(E{row}:{get_column_letter(total_days + 10)}{row})'
        sheet[f'{get_column_letter(total_days + 13)}{row}'].value = f'=ROUND({get_column_letter(total_days + 12)}{row}/{0.01 * total_days}, 1)'

# Main function
def main():
    workbook = Workbook(write_only=True)  # Open in write-only mode
    year = datetime.now().year
    workbook.remove(workbook.active)  # Remove default sheet

    workers = ['Sheriff Akatugba', 'Bassey Nton Nton', 'Denis Esikpong', 'Obiora Anointing Chimnadindu', ...]

    total_days_dict = {}

    # Create sheets for each month
    for month in list(calendar.month_name)[1:]:
        sheet, avg_col, remarks_col, total_days = create_month_sheet(workbook, month, year)
        total_days_dict[month] = total_days

        # Add worker names
        for i, worker in enumerate(workers, start=19):
            sheet.cell(row=i, column=4).value = worker

        # Apply data validation and add formulas
        apply_data_validation(sheet, 19, len(workers) + 19, 5, total_days + 5)
        add_formulas(sheet, total_days, workers)

    # Add attendance summary
    add_summary(workbook, workers, total_days_dict)

    # Add Charts
    plot_charts(workbook, workers, total_days_dict)

    # Add formatting
    format_sheet(workbook, workers, total_days_dict)

    # Add size formatting
    format_cell_size(workbook, workers,total_days_dict)

    # Save the workbook
    workbook.save("NR_ATTENDANCE_PROTOTYPE.xlsx")

if __name__ == "__main__":
    main()
