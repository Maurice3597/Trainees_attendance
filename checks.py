from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# Load the existing workbook or create a new one
workbook = Workbook()  # Replace with your filename
sheet = workbook.create_sheet("unconditional")
workbook.remove(workbook.active)  # Or specify the sheet by name

ATTENDANCE_OPTIONS = ["P", "L", "E", "A","N"]
for i in ATTENDANCE_OPTIONS:
    def format_cell_size(workbook, workers, total_days_dict):
        for sheet in workbook.worksheets:
            td = total_days_dict[sheet.title]
            omit_cols = list([col_let(td + x) for x in range(5, 12, 6)])

            sheet.column_dimensions["D"].width = 36
            sheet.column_dimensions['A'] = 5

            for i in range(td + 5, td + 18, 6):
                sheet.column_dimensions[col_let(i)].width = 3

            sheet.row_dimensions[1].height = 5
            sheet.row_dimensions[2].height = 12
            sheet.row_dimensions[3].height = 5
            sheet.row_dimensions[16].height = 5

            for col in range(5, td + 12):
                if col_let(col) not in omit_cols:
                    sheet.column_dimensions[col_let(col)].width = 5

            for row in range(17, 101):
                sheet.row_dimensions[2].height = 5

            # sheet['B2'].font = Font(name='Times New Roman', size=1, bold=True, color='ffffff')
            # sheet['B2'].value = 'Gocity Group Developers / Engineers Attendance Record'




def add_image_to_excel(image_path, cell_location, excel_file):
    """
    Adds an image to a specified cell in an Excel file using openpyxl.

    Args:
        image_path (str): The path to the image file.
        cell_location (str): The cell location (e.g., "A1") where the image should be placed.
        excel_file (str): The name of the Excel file to create or update.
    """

    try:
        # Load the workbook if it exists, otherwise create a new one
        workbook = Workbook()  # Create a new workbook by default.  If you want to load, use load_workbook()
        sheet = workbook.active

        # Create an Image object
        img = Image(image_path)

        # Add the image to the cell
        sheet.add_image(img, cell_location)

        # Adjust column width and row height (optional, but often improves appearance)
        cell = sheet[cell_location]  # Get the cell object
        column_letter = cell_location[0]  # Extract the column letter (e.g., "A")
        sheet.column_dimensions[column_letter].width = 20  # Adjust as needed
        sheet.row_dimensions[cell.row].height = 70  # Adjust as needed

        # Save the workbook
        workbook.save(excel_file)

        print(f"Image added to {cell_location} in {excel_file}")

    except FileNotFoundError:
        print(f"Error: Image file not found at {image_path}")
    except Exception as e:
        print(f"An error occurred: {e}")


# Example Usage:
if __name__ == '__main__':
    image_path = "path/to/your/image.png"  # Replace with the actual path to your image
    cell_location = "B2"  # Replace with the desired cell location
    excel_file = "image_example.xlsx"

    add_image_to_excel(image_path, cell_location, excel_file)
